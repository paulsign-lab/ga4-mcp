"""
Janibell 마케팅 인프라 세팅 체크리스트 + 실행계획서 생성
구글 시트로 변환 가능한 .xlsx 파일을 만든다.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()

# ============ 컬러 팔레트 (Janibell 브랜드 컨펌 컬러 기반) ============
COLOR_PRIMARY = "1E3A2A"      # 딥 그린 (브랜드 메인)
COLOR_ACCENT = "C9A66B"        # 따뜻한 골드
COLOR_HEADER_BG = "1E3A2A"
COLOR_HEADER_FG = "FFFFFF"
COLOR_BANNER_BG = "F4EFE6"     # 크림 베이지
COLOR_ROW_ALT = "FBFAF6"
COLOR_PRIORITY_HIGH = "F2D7D5"
COLOR_PRIORITY_MID = "FCF3CF"
COLOR_PRIORITY_LOW = "D5F5E3"
COLOR_DONE = "D6EAF8"
COLOR_NOTE = "FAE5D3"
COLOR_BORDER = "9C8E73"

thin = Side(border_style="thin", color=COLOR_BORDER)
BORDER_ALL = Border(left=thin, right=thin, top=thin, bottom=thin)

FONT_HEADER = Font(name="Arial", bold=True, color=COLOR_HEADER_FG, size=11)
FONT_BODY = Font(name="Arial", size=10)
FONT_TITLE = Font(name="Arial", bold=True, size=18, color=COLOR_PRIMARY)
FONT_SUB = Font(name="Arial", bold=True, size=12, color=COLOR_PRIMARY)
FONT_NOTE = Font(name="Arial", italic=True, size=10, color="555555")

ALIGN_HEAD = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_CELL = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)

def style_header_row(ws, row, last_col):
    for c in range(1, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_HEAD
        cell.border = BORDER_ALL

def style_body_row(ws, row, last_col, alt=False):
    for c in range(1, last_col + 1):
        cell = ws.cell(row=row, column=c)
        if alt:
            cell.fill = PatternFill("solid", fgColor=COLOR_ROW_ALT)
        cell.font = FONT_BODY
        cell.alignment = ALIGN_CELL
        cell.border = BORDER_ALL

def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def add_title(ws, title, subtitle, last_col):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    c = ws.cell(row=1, column=1, value=title)
    c.font = FONT_TITLE
    c.fill = PatternFill("solid", fgColor=COLOR_BANNER_BG)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 36

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    s = ws.cell(row=2, column=2 if False else 1, value=subtitle)
    s.font = FONT_NOTE
    s.fill = PatternFill("solid", fgColor=COLOR_BANNER_BG)
    s.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 22

# ===================================================================
# 0) 시작하기 (README)
# ===================================================================
ws = wb.active
ws.title = "0_시작하기"
add_title(ws, "Janibell 마케팅 인프라 세팅 체크리스트",
          "GA4 / GTM / Pixel / Search Console / SEO / Looker Studio 통합 실행 계획서  ·  Last update: 2026-05-14",
          last_col=4)

sections = [
    ("이 파일의 목적",
     "janibell.com Shopify 스토어가 도메인 연결까지 완료된 시점에서, 마케팅 인프라(데이터 수집·SEO·리포팅) 세팅을 누락 없이 끝내기 위한 마스터 체크리스트입니다.\n\n각 시트를 순서대로 따라가며 상태(✓) 열에 체크하면 됩니다."),
    ("시트 구성",
     "1_마스터체크리스트  →  전체 작업 목록 (50+ 항목, 카테고리·우선순위·의존성·상태)\n2_GA4_이벤트설계   →  B2B에 맞춘 GA4 이벤트 매트릭스 (이벤트명·파라미터·트리거)\n3_GTM_태그매핑     →  GTM에서 만들어야 하는 태그·트리거·변수 목록\n4_Looker_위젯스펙  →  일일 B2B 리드 대시보드 1페이지 스펙\n5_SearchConsole_운영 →  Search Console 일일/주간/월간 점검 루틴\n6_SEO_페이지별     →  페이지별 title / meta description / H1 / 구조화 데이터 체크\n7_사이트_발견이슈   →  크롤링 분석 중 발견된 즉시 조치 이슈"),
    ("색상 규칙",
     "● 우선순위 P0(필수) = 빨강 배경    ● P1(중요) = 노랑 배경    ● P2(권장) = 초록 배경\n● 상태가 '완료'인 행 = 파랑 배경 (조건부 서식 자동 적용)\n● 메모/주의 = 살구색 배경"),
    ("진행 권장 순서",
     "Step 1.  GA4 속성 + GTM 컨테이너 생성 (계정·권한 셋업)\nStep 2.  GTM 컨테이너 JSON Import (janibell_gtm_container.json)\nStep 3.  Shopify theme.liquid에 GTM 스니펫 + dataLayer 코드 삽입\nStep 4.  Meta Pixel · Search Console 연결\nStep 5.  GA4 DebugView로 이벤트 점검\nStep 6.  SEO 기초 작업 (title/meta/structured data/sitemap)\nStep 7.  Looker Studio 대시보드 만들기\nStep 8.  운영 루틴 (일일·주간·월간) 체크리스트화"),
    ("연관 산출물",
     "·  Janibell_마케팅인프라_가이드.docx → 개념 설명 + 단계별 스크린샷 가이드\n·  janibell_gtm_container.json → GTM 컨테이너 Import용 파일\n·  janibell_dataLayer_snippet.liquid → theme.liquid에 붙여넣을 dataLayer 코드"),
]

set_widths(ws, [30, 90, 4, 4])
row = 4
for title, body in sections:
    ws.cell(row=row, column=1, value=title).font = FONT_SUB
    ws.cell(row=row, column=1).alignment = Alignment(vertical="top")
    c = ws.cell(row=row, column=2, value=body)
    c.font = FONT_BODY
    c.alignment = ALIGN_CELL
    ws.row_dimensions[row].height = max(60, body.count("\n") * 18 + 30)
    row += 1

# ===================================================================
# 1) 마스터 체크리스트
# ===================================================================
ws = wb.create_sheet("1_마스터체크리스트")
add_title(ws,
          "Janibell 마스터 체크리스트",
          "전체 작업 50+ 항목. 우선순위·예상소요·의존성·상태를 한 눈에 확인하고 체크합니다.",
          last_col=10)

headers = ["#", "카테고리", "작업명", "세부 설명 / 산출물", "담당", "우선순위(P0/P1/P2)", "예상 소요", "의존성(#)", "상태", "가이드 섹션"]
for i, h in enumerate(headers, start=1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, len(headers))

tasks = [
    # 사전 준비
    ("A. 사전준비", "Google 계정 정리", "janibell 운영용 Google 계정 1개를 마스터로 지정. 개인계정 X. (예: ops@magikan.com)", "사장님/대표", "P0", "10분", "", "1.1"),
    ("A. 사전준비", "팀원 권한 정책 결정", "GA4/GTM/Search Console에 누가 어떤 권한(관리자/편집자/조회)을 가질지 표로 정리", "사장님/대표", "P0", "20분", "", "1.1"),
    ("A. 사전준비", "도메인 소유권 확인 방법 결정", "DNS TXT 추가 vs HTML 파일 업로드 vs GA4 태그 인증 중 선택. Shopify는 DNS TXT 권장.", "개발사", "P0", "10분", "", "1.2"),

    # GA4
    ("B. GA4", "GA4 계정·속성 생성", "analytics.google.com → 계정 'Janibell' → 속성 'Janibell - janibell.com'. 시간대 KST, 통화 USD", "마케팅", "P0", "20분", "1", "2.1"),
    ("B. GA4", "데이터 스트림(웹) 생성", "웹 스트림 'janibell.com' 추가 → 측정 ID(G-XXXXXXX) 확보. 향상된 측정 ON", "마케팅", "P0", "10분", "4", "2.2"),
    ("B. GA4", "데이터 보존 14개월 설정", "관리 → 데이터 보존 → 사용자 및 이벤트 데이터 보존을 14개월로", "마케팅", "P1", "5분", "4", "2.3"),
    ("B. GA4", "내부 트래픽 제외 IP 등록", "사무실/대표/개발사 IP를 internal로 태깅. 데이터 필터 'Internal Traffic'을 활성", "마케팅", "P1", "15분", "4", "2.4"),
    ("B. GA4", "사용자 IP 익명화/지역 보고서 확인", "유럽 트래픽 대응. 기본 설정 그대로 OK", "마케팅", "P2", "5분", "4", "2.5"),
    ("B. GA4", "전환 이벤트 8개 마킹", "request_quote / contact_form_submit / catalog_download_complete / file_download / newsletter_subscribe / phone_click / email_click / klaviyo_form_submit → '전환'으로 표시", "마케팅", "P0", "10분", "4,18", "2.6"),
    ("B. GA4", "맞춤 측정기준(Custom Dimension) 등록", "industry / inquiry_type / product_handle / product_category / file_name / form_location 6개를 이벤트 범위로 등록", "마케팅", "P0", "20분", "4,18", "2.7"),
    ("B. GA4", "Google Ads 연결(미사용 시 보류)", "광고 계정이 생기면 연결. 현재는 SKIP 가능", "마케팅", "P2", "10분", "4", "2.8"),
    ("B. GA4", "Search Console 연결(GA4↔GSC)", "관리 → Search Console 연결. 인수전 메뉴에 노출", "마케팅", "P1", "10분", "4,21", "2.9"),

    # GTM
    ("C. GTM", "GTM 계정·컨테이너 생성", "tagmanager.google.com → 계정 'Janibell' → 컨테이너 'janibell.com' (웹)", "마케팅", "P0", "10분", "1", "3.1"),
    ("C. GTM", "GTM 컨테이너 ID(GTM-XXXX) 확보", "설치 코드 2개(head/body) 복사. theme.liquid에 삽입할 때 사용", "마케팅", "P0", "5분", "13", "3.2"),
    ("C. GTM", "Shopify theme.liquid에 GTM 코드 삽입", "<head> 직후, <body> 직후 두 군데. 'janibell_dataLayer_snippet.liquid' 함께 적용", "개발사", "P0", "20분", "13,14", "3.3"),
    ("C. GTM", "GTM 컨테이너 JSON Import", "janibell_gtm_container.json → GTM 관리 → 컨테이너 가져오기 → '병합 + 기존 항목 덮어쓰기'", "마케팅", "P0", "10분", "13", "3.4"),
    ("C. GTM", "GA4 Configuration 태그 한 개만 유지", "GA4-Config 태그가 'All Pages'에 1번만 발동되는지 확인 (중복 발동 금지)", "마케팅", "P0", "10분", "16", "3.5"),
    ("C. GTM", "Preview 모드로 이벤트 19개 전수 점검", "GTM 미리보기 → 실제 페이지에서 클릭/제출하며 dataLayer push 검증", "마케팅", "P0", "60분", "15,16", "3.6"),
    ("C. GTM", "Container Publish (Version v1.0)", "검증 완료 후 Submit → 버전명 'v1.0 - Initial launch' + 메모 작성", "마케팅", "P0", "5분", "18", "3.7"),

    # dataLayer / 이벤트 코드
    ("D. dataLayer", "전 페이지 공통 dataLayer 변수 push", "user_logged_in / page_type / industry(컬렉션별) 자동 push", "개발사", "P0", "20분", "15", "4.1"),
    ("D. dataLayer", "view_item / view_item_list 자동 발동", "PDP/PLP에서 product_handle, product_title push", "개발사", "P0", "30분", "15", "4.2"),
    ("D. dataLayer", "Catalog 다운로드 이벤트 연결", "Klaviyo 폼 submit 후 PDF 다운로드 시 catalog_download_complete push", "개발사", "P0", "30분", "15", "4.3"),
    ("D. dataLayer", "Contact 폼 'inquiry_type' 라디오 추가", "현재 기본 폼(Name/Email/Message)만 있음. industry · inquiry_type · company · quantity 필드 추가", "개발사", "P0", "1시간", "", "4.4"),
    ("D. dataLayer", "전화/이메일 클릭 자동 캡처", "tel:, mailto: 링크 클릭 시 phone_click / email_click 발동", "마케팅", "P1", "10분", "16", "4.5"),
    ("D. dataLayer", "파일 다운로드(PDF/매뉴얼) 자동 캡처", "/files/*.pdf 클릭 시 file_download + file_name 파라미터", "마케팅", "P1", "10분", "16", "4.6"),

    # Meta Pixel / SNS
    ("E. Pixel", "Meta(Facebook) 비즈니스 계정 정리", "business.facebook.com → 광고 계정 + Pixel 1개 생성", "마케팅", "P1", "30분", "1", "5.1"),
    ("E. Pixel", "Meta Pixel ID 확보 후 GTM 태그 연결", "GTM 'Meta Pixel - Base' 태그에 ID 입력 후 활성. PageView 트리거 'All Pages'", "마케팅", "P1", "10분", "26", "5.2"),
    ("E. Pixel", "Pixel 이벤트 매핑", "Lead = request_quote · CompleteRegistration = catalog_download_complete · Contact = contact_form_submit", "마케팅", "P1", "20분", "26", "5.3"),
    ("E. Pixel", "Conversion API(CAPI) 도입 검토", "iOS14.5+ 추적 손실 대응. 향후 광고 집행 시 적용 검토", "마케팅", "P2", "보류", "27", "5.4"),

    # Search Console / SEO
    ("F. SearchConsole", "Search Console 속성 추가(도메인 속성)", "search.google.com/search-console → '도메인' 속성으로 janibell.com. DNS TXT 인증", "마케팅", "P0", "15분", "3", "6.1"),
    ("F. SearchConsole", "sitemap.xml 제출", "Shopify는 /sitemap.xml 자동 생성. Search Console에서 제출", "마케팅", "P0", "5분", "30", "6.2"),
    ("F. SearchConsole", "Bing Webmaster Tools 연동(선택)", "Search Console에서 Bing으로 1-click 가져오기. 미국 트래픽 중 Bing 비중 약 7% 대응", "마케팅", "P2", "10분", "30", "6.3"),
    ("F. SearchConsole", "URL 검사 + 'Google에 색인 요청' 5개 핵심 페이지", "홈 / Brand Story / Sustainability / Contact / Catalogue 우선 색인", "마케팅", "P1", "10분", "30", "6.4"),
    ("F. SearchConsole", "GA4와 Search Console 양방향 연결", "GA4 관리 → Search Console 연결 + Search Console 설정 → 사용자 추가에 GA4 계정", "마케팅", "P1", "10분", "12,30", "6.5"),

    # SEO 작업
    ("G. SEO기초", "Title/Meta description 페이지별 점검·작성", "홈 / 카테고리 8개 / Technology / Brand Story / Sustainability / FAQ / Manuals / Contact / Catalogue", "마케팅+카피", "P0", "3시간", "30", "7.1"),
    ("G. SEO기초", "구조화 데이터 Organization/Product/Breadcrumb 추가", "JSON-LD를 theme.liquid · product.liquid에 삽입", "개발사", "P0", "2시간", "15", "7.2"),
    ("G. SEO기초", "Open Graph + Twitter Card 메타 점검", "공유 시 미리보기 이미지·제목·설명 일관성 확인", "마케팅", "P1", "1시간", "30", "7.3"),
    ("G. SEO기초", "robots.txt 점검", "Shopify 기본 robots.txt 확인. /collections/*사본 경로는 disallow 권장", "개발사", "P1", "10분", "", "7.4"),
    ("G. SEO기초", "Canonical URL 정리(중복 collection 정리)", "/collections/home사본 등 사본 컬렉션 7개 삭제 또는 redirect", "개발사", "P0", "30분", "", "7.5"),
    ("G. SEO기초", "이미지 alt 속성 일괄 점검", "PDP 메인 이미지 + Hero 영역 alt 비어있는 항목 채우기", "마케팅+카피", "P1", "2시간", "", "7.6"),
    ("G. SEO기초", "Page speed (Core Web Vitals) 측정", "PageSpeed Insights로 홈/PDP 각각 측정 → LCP/CLS 점검", "개발사", "P1", "30분", "", "7.7"),

    # 검증
    ("H. 검증/QA", "GA4 DebugView로 이벤트 19개 발동 확인", "Chrome 확장 'GA Debugger' ON → 모든 이벤트 한 번씩 발동 → DebugView에서 파라미터 검증", "마케팅", "P0", "60분", "19", "8.1"),
    ("H. 검증/QA", "Tag Assistant로 GTM 발동 점검", "tagassistant.google.com → janibell.com 미리보기 세션 → 태그 발동 카운트 확인", "마케팅", "P0", "30분", "19", "8.2"),
    ("H. 검증/QA", "Real-time 보고서로 트래픽 흐름 확인", "GA4 실시간 보고서에서 본인 세션 추적", "마케팅", "P0", "10분", "19", "8.3"),

    # Looker Studio
    ("I. Looker", "Looker Studio 보고서 생성(B2B 리드 대시보드)", "lookerstudio.google.com → 빈 보고서 → GA4 데이터 소스 연결", "마케팅", "P0", "20분", "4,18", "9.1"),
    ("I. Looker", "1페이지 위젯 12개 배치", "4_Looker_위젯스펙 시트 참고. 14개 KPI/차트", "마케팅", "P0", "2시간", "44", "9.2"),
    ("I. Looker", "일일 이메일 스케줄(매일 09:00 KST) 설정", "보고서 → 이메일 일정 → 평일 09:00 / PDF / 본인+대표", "마케팅", "P1", "10분", "45", "9.3"),

    # 운영 루틴
    ("J. 운영루틴", "일일 점검 5분 루틴 문서화", "5_SearchConsole_운영 시트의 일일 항목 6개를 매일 09:30 체크", "마케팅", "P0", "30분", "", "10.1"),
    ("J. 운영루틴", "주간 리뷰 30분 회의 세팅", "매주 월 11:00. Looker 대시보드 1페이지 + 인사이트 슬랙 공유", "마케팅", "P1", "30분", "44", "10.2"),
    ("J. 운영루틴", "월간 SEO 리뷰 1시간", "Search Console 'Performance' → 상위 쿼리·랜딩페이지·CTR 변화", "마케팅", "P1", "1시간", "30", "10.3"),
]

start = 5
for idx, (cat, name, desc, owner, prio, eta, dep, sec) in enumerate(tasks, start=1):
    row = start + idx - 1
    ws.cell(row=row, column=1, value=idx)
    ws.cell(row=row, column=2, value=cat)
    ws.cell(row=row, column=3, value=name)
    ws.cell(row=row, column=4, value=desc)
    ws.cell(row=row, column=5, value=owner)
    ws.cell(row=row, column=6, value=prio)
    ws.cell(row=row, column=7, value=eta)
    ws.cell(row=row, column=8, value=dep)
    ws.cell(row=row, column=9, value="대기")
    ws.cell(row=row, column=10, value=sec)
    style_body_row(ws, row, 10, alt=(idx % 2 == 0))
    ws.row_dimensions[row].height = max(36, len(desc) // 40 * 14 + 28)

# 상태 dropdown
dv_status = DataValidation(type="list", formula1='"대기,진행중,완료,보류,제외"', allow_blank=True)
dv_status.add(f"I{start}:I{start+len(tasks)-1}")
ws.add_data_validation(dv_status)

# 조건부 서식: 우선순위
ws.conditional_formatting.add(f"F{start}:F{start+len(tasks)-1}",
    CellIsRule(operator="equal", formula=['"P0"'], fill=PatternFill("solid", fgColor=COLOR_PRIORITY_HIGH)))
ws.conditional_formatting.add(f"F{start}:F{start+len(tasks)-1}",
    CellIsRule(operator="equal", formula=['"P1"'], fill=PatternFill("solid", fgColor=COLOR_PRIORITY_MID)))
ws.conditional_formatting.add(f"F{start}:F{start+len(tasks)-1}",
    CellIsRule(operator="equal", formula=['"P2"'], fill=PatternFill("solid", fgColor=COLOR_PRIORITY_LOW)))

# 조건부 서식: 상태 완료
ws.conditional_formatting.add(f"A{start}:J{start+len(tasks)-1}",
    FormulaRule(formula=[f'$I{start}="완료"'], fill=PatternFill("solid", fgColor=COLOR_DONE)))

set_widths(ws, [5, 14, 32, 60, 12, 16, 10, 12, 12, 10])
ws.freeze_panes = "A5"

# ===================================================================
# 2) GA4 이벤트 설계 (B2B)
# ===================================================================
ws = wb.create_sheet("2_GA4_이벤트설계")
add_title(ws,
          "Janibell GA4 이벤트 설계 (B2B 리드 중심)",
          "Shopify 기본 이커머스 + B2B 견적·카탈로그 다운로드 중심으로 19개 이벤트를 설계.",
          last_col=10)

headers = ["#", "이벤트명", "분류(GA4 기본/추천/커스텀)", "B2B 우선순위", "전환 표시", "트리거 조건", "트리거 페이지", "필수 파라미터", "사용 케이스 / 보고서", "비고"]
for i, h in enumerate(headers, start=1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, len(headers))

events = [
    ("page_view", "기본", "필수", "X", "모든 페이지 로드", "전체", "page_location, page_title, page_type", "전체 트래픽·랜딩·이탈 측정", "GA4 자동 수집"),
    ("session_start", "기본", "필수", "X", "세션 시작", "전체", "(자동)", "세션 수·신규/재방문", "자동"),
    ("scroll", "향상된 측정", "참고", "X", "페이지 90% 스크롤", "전체", "(자동)", "콘텐츠 소비도", "향상된 측정 ON"),
    ("click(outbound)", "향상된 측정", "참고", "X", "외부 도메인 링크 클릭", "전체", "link_url, link_domain", "외부 유출 추적 (magikan.com 등)", "향상된 측정 ON"),
    ("view_item_list", "추천(ecommerce)", "참고", "X", "PLP 로드", "/collections/*", "item_list_name(=industry), items[]", "카테고리(HOME/BABY/PET/MEDICAL 등) 인기도", "B2B 데이터 설계의 핵심"),
    ("view_item", "추천(ecommerce)", "참고", "X", "PDP 로드", "/products/*", "items[](item_id=product_handle, item_name, item_category=industry, item_brand=Janibell)", "어느 제품이 가장 많이 조회되었는지", "전환 대비 깔때기 상단"),
    ("select_item", "추천(ecommerce)", "참고", "X", "PLP 카드 클릭", "/collections/*", "items[]", "PLP → PDP 전환율", "GTM 변수 활용"),
    ("nav_click", "커스텀", "참고", "X", "GNB 메뉴 클릭", "전체", "nav_label, nav_section(top/footer), nav_target", "Industry별 메뉴 클릭 분포", "B2B 산업군 관심도 판별"),
    ("request_quote", "커스텀", "최우선", "O", "PDP 'Contact Us / Request Quote' 버튼 클릭", "/products/*", "product_handle, product_title, industry, source_section", "PDP→견적 전환율, 제품별 인기 견적", "B2B 핵심 KPI"),
    ("contact_form_view", "커스텀", "필수", "X", "Contact 페이지 로드", "/pages/contact-us", "form_location, referrer_page", "Contact 페이지 도달 분석", "퍼널 상단"),
    ("contact_form_start", "커스텀", "필수", "X", "Contact 폼 첫 필드 focus", "/pages/contact-us", "form_location", "폼 시작률(=중도이탈 측정)", "퍼널 중간"),
    ("contact_form_submit", "커스텀", "최우선", "O", "Contact 폼 정상 제출", "/pages/contact-us, /pages/*", "form_location, industry, inquiry_type(quote/sample/support/partnership), company, country", "최종 리드 수, 산업별·문의유형별 리드", "B2B 핵심 KPI"),
    ("catalog_view", "커스텀", "필수", "X", "Catalogue 페이지 로드", "/pages/catalogue", "page_location, referrer_page", "카탈로그 게이트 도달", "퍼널 상단"),
    ("catalog_download_start", "커스텀", "중요", "X", "카탈로그 다운로드 Klaviyo 폼 submit", "/pages/catalogue", "email_provided=true, industry", "Klaviyo 폼 제출 = 잠재 리드", "Klaviyo와 양방향 매칭"),
    ("catalog_download_complete", "커스텀", "최우선", "O", "PDF 다운로드 성공", "/pages/catalogue", "file_name(=catalog_2026_en.pdf), industry", "실제 카탈로그 다운수 = 콜드 리드 정의", "B2B 핵심 KPI"),
    ("file_download", "향상된 측정", "필수", "O", "/files/*.pdf 또는 /cdn/shop/files/*.pdf 클릭", "전체", "file_name, file_extension, file_path", "매뉴얼·스펙시트 등 자산 다운", "리드 스코어링 신호"),
    ("phone_click", "커스텀", "중요", "O", "tel: 링크 클릭", "전체", "phone_number, page_location", "전화 문의 시도 수", "오프라인 전환 신호"),
    ("email_click", "커스텀", "중요", "O", "mailto: 링크 클릭", "전체", "email_address, page_location", "이메일 문의 시도 수", "오프라인 전환 신호"),
    ("newsletter_subscribe", "추천", "중요", "O", "Footer Newsletter 폼 제출", "전체", "form_location=footer, source_page", "마케팅 동의 리드 수집", "리드 너처링 인풋"),
    ("klaviyo_form_submit", "커스텀", "중요", "O", "Klaviyo 폼(팝업/임베드) 제출", "전체", "klaviyo_form_id, source_page, industry", "Klaviyo 이벤트와 GA4 동기화 검증", "Catalog 게이트와 별도"),
    ("video_play", "커스텀", "참고", "X", "Hero/PDP 영상 재생 30% 도달", "홈/PDP/Technology", "video_title, video_provider, video_percent", "브랜드 영상 소비도", "Veo3 영상 효율 측정"),
    ("manual_view", "커스텀", "참고", "X", "Manuals 페이지 PDF 클릭 직전", "/pages/manual", "product_model, language", "어느 모델 매뉴얼이 많이 열리는지", "After-sales 인사이트"),
]

for idx, e in enumerate(events, start=1):
    row = 5 + idx - 1
    ws.cell(row=row, column=1, value=idx)
    for i, v in enumerate(e, start=2):
        ws.cell(row=row, column=i, value=v)
    style_body_row(ws, row, len(headers), alt=(idx % 2 == 0))
    ws.row_dimensions[row].height = max(36, len(e[6]) // 30 * 14 + 28)

# Conditional fmt B2B 우선순위
ws.conditional_formatting.add(f"D5:D{4+len(events)}",
    CellIsRule(operator="equal", formula=['"최우선"'], fill=PatternFill("solid", fgColor=COLOR_PRIORITY_HIGH)))
ws.conditional_formatting.add(f"D5:D{4+len(events)}",
    CellIsRule(operator="equal", formula=['"필수"'], fill=PatternFill("solid", fgColor=COLOR_PRIORITY_MID)))
ws.conditional_formatting.add(f"D5:D{4+len(events)}",
    CellIsRule(operator="equal", formula=['"중요"'], fill=PatternFill("solid", fgColor=COLOR_PRIORITY_MID)))
ws.conditional_formatting.add(f"E5:E{4+len(events)}",
    CellIsRule(operator="equal", formula=['"O"'], fill=PatternFill("solid", fgColor=COLOR_DONE)))

set_widths(ws, [5, 26, 22, 12, 10, 36, 24, 50, 50, 28])
ws.freeze_panes = "A5"

# ===================================================================
# 3) GTM 태그 매핑
# ===================================================================
ws = wb.create_sheet("3_GTM_태그매핑")
add_title(ws,
          "GTM 태그·트리거·변수 매핑표",
          "이 표 그대로 GTM 컨테이너에 만들어야 할 항목. janibell_gtm_container.json을 Import하면 자동 생성.",
          last_col=8)

headers = ["#", "구분", "이름", "타입", "발동 조건(트리거)", "주요 설정값", "GA4 전송 이벤트", "비고"]
for i, h in enumerate(headers, start=1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, len(headers))

gtm_rows = [
    ("태그", "GA4-Config", "Google 태그", "All Pages", "Measurement ID = {{const.GA4_ID}}, send_page_view=true", "(컨피그)", "유일하게 1개만 존재해야 함"),
    ("태그", "GA4-Event-view_item_list", "Google Analytics: GA4 이벤트", "PLP Page View", "event_name=view_item_list, items={{dl.items}}", "view_item_list", "dataLayer.push 기반"),
    ("태그", "GA4-Event-view_item", "GA4 이벤트", "PDP Page View", "event_name=view_item, items={{dl.items}}", "view_item", "Shopify 메타 변수 사용"),
    ("태그", "GA4-Event-select_item", "GA4 이벤트", "PLP Card Click", "event_name=select_item, item_id={{dl.product_handle}}", "select_item", ""),
    ("태그", "GA4-Event-request_quote", "GA4 이벤트", "PDP CTA Click (Contact/Quote)", "event_name=request_quote, product_handle, industry, source_section", "request_quote", "★ B2B 핵심"),
    ("태그", "GA4-Event-nav_click", "GA4 이벤트", "Nav Link Click", "event_name=nav_click, nav_label, nav_section", "nav_click", ""),
    ("태그", "GA4-Event-contact_form_view", "GA4 이벤트", "Contact Page View", "event_name=contact_form_view, form_location, referrer_page", "contact_form_view", ""),
    ("태그", "GA4-Event-contact_form_start", "GA4 이벤트", "Contact Form Focus", "event_name=contact_form_start, form_location", "contact_form_start", ""),
    ("태그", "GA4-Event-contact_form_submit", "GA4 이벤트", "Contact Form Submit (DOM 또는 dataLayer)", "event_name=contact_form_submit, industry, inquiry_type, company, country", "contact_form_submit", "★ B2B 핵심"),
    ("태그", "GA4-Event-catalog_download_start", "GA4 이벤트", "Klaviyo 폼 Submit (Catalog 게이트)", "event_name=catalog_download_start, industry", "catalog_download_start", "Klaviyo 폼 ID 트리거"),
    ("태그", "GA4-Event-catalog_download_complete", "GA4 이벤트", "PDF 다운로드 성공 콜백", "event_name=catalog_download_complete, file_name", "catalog_download_complete", "★ B2B 핵심"),
    ("태그", "GA4-Event-file_download", "GA4 이벤트", "PDF 다운로드 클릭(기타)", "event_name=file_download, file_name, file_extension", "file_download", ""),
    ("태그", "GA4-Event-phone_click", "GA4 이벤트", "tel: 클릭", "event_name=phone_click, phone_number", "phone_click", ""),
    ("태그", "GA4-Event-email_click", "GA4 이벤트", "mailto: 클릭", "event_name=email_click, email_address", "email_click", ""),
    ("태그", "GA4-Event-newsletter_subscribe", "GA4 이벤트", "Footer Newsletter Submit", "event_name=newsletter_subscribe, form_location=footer", "newsletter_subscribe", ""),
    ("태그", "GA4-Event-klaviyo_form_submit", "GA4 이벤트", "Klaviyo 일반 폼 Submit (Catalog 외)", "event_name=klaviyo_form_submit, klaviyo_form_id", "klaviyo_form_submit", ""),
    ("태그", "GA4-Event-video_play", "GA4 이벤트", "YouTube 30% 재생", "event_name=video_play, video_title, video_percent", "video_play", "Built-in YouTube 변수"),
    ("태그", "Meta-Pixel-Base", "맞춤 HTML", "All Pages", "fbq('init', PIXEL_ID); fbq('track','PageView')", "(자동 PageView)", "Pixel ID 후속 입력"),
    ("태그", "Meta-Pixel-Lead", "맞춤 HTML", "Contact Submit / Quote Request", "fbq('track','Lead', {content_name})", "(Pixel용 Lead)", ""),
    ("태그", "Meta-Pixel-CompleteReg", "맞춤 HTML", "catalog_download_complete", "fbq('track','CompleteRegistration')", "(Pixel)", ""),

    ("트리거", "All Pages", "페이지뷰", "모든 페이지", "(기본)", "", ""),
    ("트리거", "PLP Page View", "페이지뷰 (DOM Ready)", "Page Path matches RegEx ^/collections/", "", "", ""),
    ("트리거", "PDP Page View", "페이지뷰", "Page Path matches RegEx ^/products/", "", "", ""),
    ("트리거", "PLP Card Click", "Link Click (Just Links)", "Click URL matches RegEx /products/", "", "", "Wait for tags 2000ms"),
    ("트리거", "PDP CTA Click (Contact/Quote)", "All Elements 클릭", "Click Classes contains 'btn-contact' OR Click Text matches 'Contact Us|Request a Quote'", "", "", ""),
    ("트리거", "Nav Link Click", "Link Click", "Click Element matches CSS Selector 'header nav a, footer a'", "", "", ""),
    ("트리거", "Contact Page View", "페이지뷰", "Page Path equals /pages/contact-us", "", "", ""),
    ("트리거", "Contact Form Focus", "Custom Event", "event = contact_form_start", "", "", "dataLayer push 기반"),
    ("트리거", "Contact Form Submit", "Custom Event", "event = contact_form_submit", "", "", "dataLayer push 기반(theme.liquid에서 발동)"),
    ("트리거", "Klaviyo Catalog Form Submit", "Custom Event", "event = catalog_download_start", "", "", "Klaviyo 콜백 이용"),
    ("트리거", "Catalog PDF Downloaded", "Custom Event", "event = catalog_download_complete", "", "", ""),
    ("트리거", "PDF Click", "Link Click", "Click URL matches RegEx \\.pdf($|\\?)", "", "", ""),
    ("트리거", "Phone Click", "Link Click", "Click URL starts with 'tel:'", "", "", ""),
    ("트리거", "Email Click", "Link Click", "Click URL starts with 'mailto:'", "", "", ""),
    ("트리거", "Newsletter Footer Submit", "Form Submission", "Form ID equals 'newsletter-footer'", "", "", ""),
    ("트리거", "Klaviyo Generic Submit", "Custom Event", "event = klaviyo_form_submit", "", "", ""),
    ("트리거", "YouTube 30% Play", "YouTube 동영상", "Capture: Start/Complete/Progress(25,50,75,90)", "", "", "Built-in YouTube"),

    ("변수", "const.GA4_ID", "상수", "—", "G-XXXXXXX", "—", "GA4 측정 ID 1회만 변경"),
    ("변수", "dl.items", "데이터영역 변수", "—", "items", "—", "view_item/list 페이로드"),
    ("변수", "dl.product_handle", "데이터영역 변수", "—", "product_handle", "—", ""),
    ("변수", "dl.product_title", "데이터영역 변수", "—", "product_title", "—", ""),
    ("변수", "dl.industry", "데이터영역 변수", "—", "industry", "—", "HOME/BABY/PETS/CARE/MEDICAL/PUBLIC/HOSPITALITY/OFFICE"),
    ("변수", "dl.inquiry_type", "데이터영역 변수", "—", "inquiry_type", "—", "quote/sample/support/partnership"),
    ("변수", "dl.form_location", "데이터영역 변수", "—", "form_location", "—", "header/footer/contact-page/pdp"),
    ("변수", "dl.file_name", "데이터영역 변수", "—", "file_name", "—", ""),
    ("변수", "js.fileExtension", "JavaScript", "—", "Click URL에서 확장자 추출", "—", ""),
    ("변수", "lookup.pagePath_industry", "Lookup Table", "—", "Path → industry 매핑", "—", "/collections/baby → BABY 등"),
]

for idx, r in enumerate(gtm_rows, start=1):
    row = 5 + idx - 1
    ws.cell(row=row, column=1, value=idx)
    for i, v in enumerate(r, start=2):
        ws.cell(row=row, column=i, value=v)
    style_body_row(ws, row, len(headers), alt=(idx % 2 == 0))
    ws.row_dimensions[row].height = 30

ws.conditional_formatting.add(f"B5:B{4+len(gtm_rows)}",
    CellIsRule(operator="equal", formula=['"태그"'], fill=PatternFill("solid", fgColor="EAF2F8")))
ws.conditional_formatting.add(f"B5:B{4+len(gtm_rows)}",
    CellIsRule(operator="equal", formula=['"트리거"'], fill=PatternFill("solid", fgColor="FEF9E7")))
ws.conditional_formatting.add(f"B5:B{4+len(gtm_rows)}",
    CellIsRule(operator="equal", formula=['"변수"'], fill=PatternFill("solid", fgColor="EAFAF1")))

set_widths(ws, [5, 8, 32, 22, 48, 48, 22, 30])
ws.freeze_panes = "A5"

# ===================================================================
# 4) Looker Studio 위젯 스펙
# ===================================================================
ws = wb.create_sheet("4_Looker_위젯스펙")
add_title(ws,
          "Looker Studio - Janibell B2B 리드 1페이지 대시보드 스펙",
          "GA4 데이터 소스 기준. 위에서 아래로 6행, 좌→우 4컬럼 그리드. 행 1~4는 KPI 카드, 행 5~6은 차트.",
          last_col=8)

headers = ["#", "섹션", "위젯명", "시각화 타입", "차원(Dimension)", "측정항목(Metric)", "필터/조건", "비고"]
for i, h in enumerate(headers, start=1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, len(headers))

widgets = [
    ("Row 1 - 핵심 KPI", "어제 리드 수", "스코어카드", "—", "이벤트수 (이벤트 IN (request_quote, contact_form_submit, catalog_download_complete))", "날짜 = 어제", "비교: 전일 대비"),
    ("Row 1", "이번주 누적 리드", "스코어카드", "—", "이벤트수 (이벤트 IN 위와 동일)", "기간 = This Week", "비교: 지난주"),
    ("Row 1", "이번달 누적 리드", "스코어카드", "—", "이벤트수 (이벤트 IN 위와 동일)", "기간 = This Month", "비교: 지난달"),
    ("Row 1", "리드당 세션수(효율)", "스코어카드", "—", "세션수 / 리드 이벤트수", "기간 = This Month", "낮을수록 좋음 (정량적 효율 지표)"),
    ("Row 2 - 깔때기", "리드 깔때기", "막대 그래프(가로)", "이벤트명", "이벤트수", "이벤트 IN (page_view, view_item, request_quote, contact_form_submit, catalog_download_complete)", "퍼널 5단계 비주얼"),
    ("Row 2", "리드 유형 분포", "도넛 차트", "inquiry_type", "이벤트수", "이벤트=contact_form_submit", "quote/sample/support/partnership 비중"),
    ("Row 3 - 산업군", "산업군별 리드", "수평 막대", "industry (커스텀 측정기준)", "리드 이벤트수", "리드 이벤트 IN 위 3개", "HOME/BABY/PETS/CARE/MEDICAL/PUBLIC/HOSPITALITY/OFFICE"),
    ("Row 3", "산업군별 PDP 조회", "수평 막대", "industry", "view_item 이벤트수", "이벤트=view_item", "PDP 인기 vs 실제 전환 차이 비교용"),
    ("Row 4 - 유입", "유입 채널 TOP 10", "표", "session_source / medium", "세션수, 리드수, 리드율(=리드수/세션수)", "기간 = This Month", "Direct/Organic/Email/Referral 등"),
    ("Row 4", "랜딩페이지 TOP 10", "표", "랜딩페이지", "세션수, 리드수", "기간 = This Month", "어떤 페이지로 들어와야 리드가 되는가"),
    ("Row 5 - 콘텐츠", "제품(상품) 인기 TOP 10", "표", "product_handle(=item_id)", "view_item 이벤트수, request_quote 이벤트수", "이벤트 IN (view_item, request_quote)", "PDP→견적 전환 계산"),
    ("Row 5", "다운로드된 자산 TOP 5", "표", "file_name", "file_download 이벤트수", "이벤트=file_download OR catalog_download_complete", "마케팅 자산 효율"),
    ("Row 6 - 트렌드", "일별 리드 추이(30일)", "꺾은선", "날짜", "리드 이벤트수", "최근 30일", "주말 vs 평일 흐름 확인"),
    ("Row 6", "검색어/유입 키워드 TOP 10", "표", "Search Console: 검색어", "노출수, 클릭수, CTR, 평균순위", "GSC 연결 후", "SEO 인사이트"),
]

for idx, w in enumerate(widgets, start=1):
    row = 5 + idx - 1
    ws.cell(row=row, column=1, value=idx)
    for i, v in enumerate(w, start=2):
        ws.cell(row=row, column=i, value=v)
    style_body_row(ws, row, len(headers), alt=(idx % 2 == 0))
    ws.row_dimensions[row].height = 38

set_widths(ws, [5, 18, 28, 18, 28, 36, 32, 28])
ws.freeze_panes = "A5"

# ===================================================================
# 5) Search Console 운영 루틴
# ===================================================================
ws = wb.create_sheet("5_SearchConsole_운영")
add_title(ws,
          "Search Console 운영 체크리스트",
          "일일 5분 · 주간 30분 · 월간 60분 루틴. 이 시트만 따라 가도 SEO 건강도가 유지됩니다.",
          last_col=6)

headers = ["#", "주기", "메뉴 (Search Console)", "체크 항목", "정상 범위 / 기준", "이상 시 액션"]
for i, h in enumerate(headers, start=1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, len(headers))

gsc_items = [
    ("일일", "개요 (Overview)", "어제 총 클릭수·노출수", "전일 대비 ±20% 이내", "급감 시: Performance 화면에서 어떤 페이지/쿼리에서 빠졌는지 확인"),
    ("일일", "색인 → 페이지", "에러(Not indexed) 페이지 증가 여부", "신규 에러 0", "URL 검사 → Live Test → 'Request Indexing' 또는 코드/리다이렉트 수정"),
    ("일일", "보안 및 수동 조치", "수동 조치 / 보안 문제 알림", "0건", "발생 시 즉시 대응. 메일 알림으로 인지"),
    ("일일", "URL 검사", "신규 생성 페이지 1건 색인 요청", "—", "Janibell 신규 페이지/블로그 발행 시 'Request Indexing' 1회 클릭"),
    ("일일", "Performance", "리드 핵심 페이지(/pages/contact-us, /pages/catalogue, /products/*) 노출수 변화", "최근 7일 추세 안정", "급감 시 sitemap 재제출 + URL 검사"),
    ("일일", "(메일 알림)", "Search Console에서 온 신규 알림 메일 확인", "—", "메일 본문 링크 따라가 즉시 조치"),

    ("주간", "Performance", "검색어 TOP 20 (지난 7일 vs 직전 7일)", "주요 키워드(odor-sealing, diaper pail, sanitary disposal, B2B/wholesale) 순위 유지", "순위 하락 시 해당 페이지 제목/콘텐츠 보강"),
    ("주간", "Performance", "페이지 TOP 20 클릭/노출", "/, /pages/brand-story, /collections/baby, /pages/catalogue 상위 유지", "상위 페이지 이탈 시 콘텐츠 업데이트"),
    ("주간", "Performance", "국가별 클릭(미국/캐나다 비중)", "북미 70%+ ", "낮으면 hreflang/Geo 설정 검토"),
    ("주간", "Performance", "기기별 클릭(모바일/PC)", "모바일 50%+ 권장", "모바일 낮으면 Core Web Vitals 모바일 점수 확인"),
    ("주간", "Sitemaps", "sitemap.xml 상태", "성공 / 마지막 읽기 7일 이내", "에러 시 sitemap 재제출"),
    ("주간", "Enhancement → Core Web Vitals", "LCP/INP/CLS 'Poor' 페이지", "Poor URL 0개", "Lighthouse 측정 → 이미지·JS 최적화 의뢰"),

    ("월간", "Performance (기간 28일/3개월 비교)", "전체 클릭·노출·CTR·평균순위 추이", "월간 클릭 증가율 +5% 이상 목표", "감소 시 콘텐츠 캘린더 보강 / 백링크 전략 수립"),
    ("월간", "Coverage / Pages", "Indexed vs Not indexed 비율", "Indexed 비율 90%+", "Not indexed 사유별 분류 → noindex/canonical/4xx/5xx 점검"),
    ("월간", "Links", "내부 링크 TOP 페이지 / 외부 도메인 TOP", "Brand Story·Sustainability 내부링크 유입 확인", "전략 페이지 내부 링크 부족 시 메뉴/푸터 보강"),
    ("월간", "Enhancement → Schema.org", "Organization/Product/Breadcrumb 인식 여부", "에러 0, 유효 100%", "JSON-LD 수정 후 URL 검사 재요청"),
    ("월간", "검색 UX → Mobile Usability", "모바일 가독성 이슈 페이지", "0건", "viewport/clickable element 간격 수정"),
    ("월간", "Removals", "삭제한 페이지(/사본 컬렉션 등) 캐시 제거", "처리 상태 = Completed", "Removal Tool로 캐시 제거 후 410/redirect 적용 확인"),
]

for idx, r in enumerate(gsc_items, start=1):
    row = 5 + idx - 1
    ws.cell(row=row, column=1, value=idx)
    for i, v in enumerate(r, start=2):
        ws.cell(row=row, column=i, value=v)
    style_body_row(ws, row, len(headers), alt=(idx % 2 == 0))
    ws.row_dimensions[row].height = 38

ws.conditional_formatting.add(f"B5:B{4+len(gsc_items)}",
    CellIsRule(operator="equal", formula=['"일일"'], fill=PatternFill("solid", fgColor=COLOR_PRIORITY_HIGH)))
ws.conditional_formatting.add(f"B5:B{4+len(gsc_items)}",
    CellIsRule(operator="equal", formula=['"주간"'], fill=PatternFill("solid", fgColor=COLOR_PRIORITY_MID)))
ws.conditional_formatting.add(f"B5:B{4+len(gsc_items)}",
    CellIsRule(operator="equal", formula=['"월간"'], fill=PatternFill("solid", fgColor=COLOR_PRIORITY_LOW)))

set_widths(ws, [5, 8, 30, 40, 30, 50])
ws.freeze_panes = "A5"

# ===================================================================
# 6) SEO 페이지별
# ===================================================================
ws = wb.create_sheet("6_SEO_페이지별")
add_title(ws,
          "페이지별 SEO 점검표",
          "각 페이지의 Title / Meta Description / H1 / 구조화 데이터 / OG / Canonical을 한 번씩 점검.",
          last_col=10)

headers = ["#", "페이지명", "URL", "Title (60자 권장)", "Meta Description (155자 권장)", "H1", "구조화 데이터(JSON-LD)", "OG 이미지", "Canonical", "상태"]
for i, h in enumerate(headers, start=1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, len(headers))

seo_pages = [
    ("홈", "/", "Janibell | Premium Odor-Sealing Disposal Systems for B2B Partners", "Janibell delivers odor-sealing diaper pails and hygiene disposal systems trusted by childcare, healthcare, and retail partners across North America. GRS-certified.", "janibell", "Organization", "필수(1200x630)", "/", "현재값 유지"),
    ("HOME 컬렉션", "/collections/home-odor-sealing-pails", "Home Odor-Sealing Pails – Janibell B2B", "Premium home-use odor-sealing pails: BiiNü, AKORD Element, Champ. Designed for North American consumer and B2B retail partners.", "Home Odor-Sealing Pails", "CollectionPage + ItemList", "권장", "Pattern A URL", "URL 변경 필요"),
    ("BABY 컬렉션", "/collections/diaper-pails", "Diaper Pails Wholesale – Janibell B2B (AKORD · BiiNü · Champ)", "Wholesale odor-sealing diaper pails for childcare centers, daycares, and retail partners. AKORD, BiiNü, Champ series with continuous-liner technology.", "Baby Diaper Pails", "CollectionPage", "권장", "Pattern A URL", "URL 변경 필요"),
    ("PETS 컬렉션", "/collections/pet-waste-pails", "Pet Waste Disposal Pails – Janibell Wholesale", "Litter Champ, AKORD Pet — odor-sealing pet waste and litter pails for pet retail, kennels, grooming centers, and shelters.", "Pet Waste Pails", "CollectionPage", "권장", "Pattern A URL", "URL 변경 필요"),
    ("CARE 컬렉션", "/collections/adult-care-disposal", "Adult Care Incontinence Disposal – Janibell Wholesale", "Discreet, odor-locking adult incontinence and senior-care disposal pails for senior living, home health, and retail.", "Adult Care Disposal", "CollectionPage", "권장", "Pattern A URL", "URL 변경 필요"),
    ("MEDICAL 컬렉션", "/collections/medical-waste-pails", "Medical Waste Pails – Janibell Clinical B2B", "Clinical-grade odor-sealing waste pails for clinics, dental offices, and outpatient care. PRIX and AKORD-Medical series.", "Medical Waste Pails", "CollectionPage", "권장", "Pattern A URL", "URL 변경 필요"),
    ("PUBLIC 컬렉션", "/collections/sanitary-disposal-bins", "Sanitary Disposal Bins for Public Restrooms – Janibell Trade", "Feminine hygiene and sanitary disposal bins for public facilities, fitness, transit, and washrooms. Modern, hands-free, odor-sealed.", "Sanitary Disposal Bins", "CollectionPage", "권장", "Pattern A URL", "URL 변경 필요"),
    ("HOSPITALITY", "/collections/hospitality-sanitary-bins", "Hospitality Sanitary Bins – Hotels · Spas · Resorts", "Discreet odor-sealing sanitary bins for hotels, resorts, spas, and hospitality washrooms. Designed to elevate guest experience.", "Hospitality Sanitary Bins", "CollectionPage", "권장", "Pattern A URL", "URL 변경 필요"),
    ("OFFICE", "/collections/office-sanitary-bins", "Office Sanitary Bins – Workplace Washroom Disposal", "Modern, low-profile sanitary bins designed for corporate restrooms, co-working, and high-traffic office washrooms.", "Office Sanitary Bins", "CollectionPage", "권장", "Pattern A URL", "URL 변경 필요"),
    ("Technology", "/pages/technology", "The Science Behind the Freshness – Janibell Technology", "Continuous Liner · Double-Lock · Auto-Sealing · Recycled Materials — four patented Janibell technologies that block odor at the source.", "The Science Behind the Freshness.", "WebPage + TechArticle(선택)", "권장", "현재값 유지", "작성 필요"),
    ("Brand Story", "/pages/brand-stroy", "Brand Story – Janibell since 1996", "From Magikan's roots in Korea to a global hygiene brand: 30 years of odor-sealing innovation.", "TBD", "AboutPage", "권장", "URL 'brand-stroy' → 'brand-story'로 수정 또는 redirect", "URL 오타 수정 필요"),
    ("Sustainability", "/pages/sustainability", "Sustainability — A Fresh Home, A Cleaner Planet", "GRS-certified recycled materials, low-energy injection molding, and 2024 K-ESG Environmental Grand Prize.", "Sustainability", "AboutPage", "필수", "현재값 유지", "작성 필요"),
    ("FAQ", "/pages/faq", "FAQ – Janibell Diaper Pails, Refills, B2B Pricing", "Answers about Janibell products, refill compatibility, MOQ, lead time, and warranty.", "FAQ", "FAQPage (★ SEO 효과 큼)", "권장", "현재값 유지", "FAQPage 스키마 필수"),
    ("Manuals", "/pages/manual", "Product Manuals – Janibell User Guides & Spec Sheets", "Download manuals for AKORD, BiiNü, Champ, PRIX series. PDF available in English.", "Manuals", "WebPage", "권장", "현재값 유지", "작성 필요"),
    ("Contact Us", "/pages/contact-us", "Contact Janibell — B2B Inquiries, Quotes, and Samples", "Talk to Janibell's North America sales team. Wholesale, OEM/ODM, and partnership inquiries welcome.", "Contact Us", "ContactPage", "권장", "현재값 유지", "작성 필요"),
    ("Catalogue", "/pages/catalogue", "Download the Janibell B2B Catalogue (PDF, 2026)", "Get our 2026 product line-card and pricing for North American B2B partners — instant download.", "Catalogue", "WebPage", "필수 (1200x630)", "현재값 유지", "작성 필요"),
    ("Heritage & Awards", "/pages/heritage-awards", "Heritage & Awards – 2024 K-ESG Environmental Grand Prize", "Patents, certifications, and awards that prove Janibell's odor-sealing leadership since 1996.", "Heritage & Awards", "AboutPage", "권장", "현재값 유지", "작성 필요"),
    ("Real Homes (Blog)", "/blogs/real-homes", "Real Homes – How Families Use Janibell", "Customer stories and home-segment use cases of Janibell odor-sealing pails.", "Real Homes", "Blog + BlogPosting(개별)", "권장", "현재값 유지", "운영 시작 시 작성"),
    ("News (Blog)", "/blogs/news", "Janibell News, B2B Insights, Industry Updates", "Latest from Janibell — product launches, partnerships, and B2B hygiene-industry insights.", "Blog", "Blog + BlogPosting", "권장", "현재값 유지", "운영 시작 시 작성"),
]

for idx, r in enumerate(seo_pages, start=1):
    row = 5 + idx - 1
    ws.cell(row=row, column=1, value=idx)
    for i, v in enumerate(r, start=2):
        ws.cell(row=row, column=i, value=v)
    style_body_row(ws, row, len(headers), alt=(idx % 2 == 0))
    ws.row_dimensions[row].height = 50

set_widths(ws, [5, 16, 30, 50, 60, 18, 26, 16, 30, 16])
ws.freeze_panes = "A5"

# ===================================================================
# 7) 사이트 발견 이슈 (크롤링)
# ===================================================================
ws = wb.create_sheet("7_사이트_발견이슈")
add_title(ws,
          "janibell.com 크롤링 분석 - 즉시 조치 이슈",
          "2026-05-14 분석 기준. 본 GA4/GTM 작업 전·중에 함께 정리해야 데이터가 깨끗하게 쌓입니다.",
          last_col=7)

headers = ["#", "구분", "발견 위치", "이슈 내용", "데이터/SEO 영향", "우선순위", "권장 조치"]
for i, h in enumerate(headers, start=1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, len(headers))

issues = [
    ("URL 정합성", "GNB / 컬렉션", "BABY/PETS/CARE/MEDICAL/PUBLIC/HOSPITALITY/OFFICE 7개 컬렉션이 /collections/baby사본 등 '사본' 접미사 URL 사용 + 메뉴-핸들 매핑 불일치(BABY 메뉴가 home사본을 가리킴)", "사용자가 공유한 링크에 한글 encoded가 포함됨 + 잘못된 산업군 데이터 수집 → SEO·UX·분석 모두 부정적", "P0", "Pattern A 키워드 URL로 재핸들링: diaper-pails / pet-waste-pails / adult-care-disposal / medical-waste-pails / sanitary-disposal-bins / hospitality-sanitary-bins / office-sanitary-bins / home-odor-sealing-pails. 9_URL_리다이렉트 시트의 매핑대로 301 redirect 설정 필수"),
    ("URL 정합성", "GNB", "Brand Story 링크가 /pages/brand-stroy (오타: stroy)", "SEO·신뢰도 손상", "P0", "Shopify Admin → Pages → handle을 'brand-story'로 수정 + Redirect 추가"),
    ("정보 구조", "GNB", "'Products / Legacy / New Technology' 메뉴 3개가 모두 /pages/technology로 동일 링크", "사용자 혼란, 메뉴 클릭 지표 신뢰도 X", "P0", "각 메뉴 별 별도 페이지/앵커 분리 (또는 메뉴 단순화)"),
    ("커머스 데이터", "PDP / PLP", "전 제품 가격이 $0.00. PDP에 'Add to Cart' 없이 'Contact Us' 만 존재", "B2B 견적 모델로 의도된 것으로 보임 → GA4도 view_item/select_item만 사용하고, 결제 전환은 추적 안 함이 옳음", "P1", "의도가 맞다면 가격 표기 → 'Contact for Pricing'으로 변경. 'Request a Quote' 버튼을 PDP에 별도 추가"),
    ("벤더 메타데이터", "Shopify product.json", "vendor가 'janibell.dev'로 일괄 표시", "SEO 메타·구조화 데이터 brand 필드에 'janibell.dev'가 노출", "P1", "Shopify Admin에서 Vendor를 'Janibell'로 일괄 변경 (Bulk edit)"),
    ("스크립트 자산", "site-wide", "GA4·GTM·Meta Pixel 미설치(Klaviyo만 있음)", "현재 데이터 수집 불가. 본 가이드의 핵심 작업", "P0", "본 체크리스트 B/C/E 카테고리 작업으로 해결"),
    ("Legacy 쿠키", "site-wide", "_fwb / __bs_imweb / _imweb_login_state 쿠키 발견(아임웹 잔여 흔적)", "기능 영향 미미하지만 보안 감사 시 지적 가능", "P2", "아임웹 잔여 스크립트 제거 점검(이미 무관 시 무시)"),
    ("Klaviyo 통합", "Catalogue 페이지", "현재 /pages/catalogue에 Klaviyo 폼/PDF 다운로드 게이트가 미연결", "B2B 핵심 리드 채널 부재", "P0", "Janibell_Catalog_Download_Integration_Manual.docx 가이드대로 폼 + PDF 트리거 설치"),
    ("Contact 폼 필드", "/pages/contact-us", "Name / Email / Message만 존재. B2B에 필요한 Company / Country / Industry / Inquiry Type / Quantity 없음", "리드 퀄리티 저하, GA4에 industry/inquiry_type 측정기준 못 채움", "P0", "Shopify 섹션 contact-form.liquid 수정해 필드 추가. dataLayer push에 동일 필드 포함"),
    ("Search Console", "—", "Search Console 속성 미등록", "색인 상태 불가시, 알림 못 받음", "P0", "본 체크리스트 F.30 작업"),
    ("Pixel/광고", "—", "Meta Pixel 미설치", "유료 광고 집행 시 리타게팅·전환 추적 불가", "P1", "본 체크리스트 E 카테고리 작업"),
    ("페이지 누락", "GNB", "Sustainability 메뉴는 있으나 페이지 본문 비어있음 (Migration Notice만 표시)", "ESG/브랜드 신뢰도 손상", "P1", "Janibell_Sustainability_페이지기획.docx 기반 콘텐츠 배포"),
]

for idx, r in enumerate(issues, start=1):
    row = 5 + idx - 1
    ws.cell(row=row, column=1, value=idx)
    for i, v in enumerate(r, start=2):
        ws.cell(row=row, column=i, value=v)
    style_body_row(ws, row, len(headers), alt=(idx % 2 == 0))
    ws.row_dimensions[row].height = 48

ws.conditional_formatting.add(f"F5:F{4+len(issues)}",
    CellIsRule(operator="equal", formula=['"P0"'], fill=PatternFill("solid", fgColor=COLOR_PRIORITY_HIGH)))
ws.conditional_formatting.add(f"F5:F{4+len(issues)}",
    CellIsRule(operator="equal", formula=['"P1"'], fill=PatternFill("solid", fgColor=COLOR_PRIORITY_MID)))
ws.conditional_formatting.add(f"F5:F{4+len(issues)}",
    CellIsRule(operator="equal", formula=['"P2"'], fill=PatternFill("solid", fgColor=COLOR_PRIORITY_LOW)))

set_widths(ws, [5, 16, 28, 50, 40, 10, 50])
ws.freeze_panes = "A5"

# ===================================================================
# 9) URL 리다이렉트 매핑 (Pattern A 적용)
# ===================================================================
ws = wb.create_sheet("9_URL_리다이렉트")
add_title(ws,
          "Shopify URL Redirect 매핑 (Pattern A 적용)",
          "Shopify Admin → Online Store → Navigation → URL redirects → 'Add URL redirect'에서 아래 표를 그대로 입력. 301 리다이렉트로 SEO 자산을 새 URL로 이동.",
          last_col=5)

headers = ["#", "From (기존 URL)", "To (새 URL - Pattern A)", "의미하는 산업군", "비고"]
for i, h in enumerate(headers, start=1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, len(headers))

redirects = [
    ("/collections/home사본", "/collections/diaper-pails", "BABY", "원래 BABY 메뉴가 가리키던 핸들"),
    ("/collections/baby사본", "/collections/pet-waste-pails", "PETS", "원래 PETS 메뉴가 가리키던 핸들"),
    ("/collections/medical사본", "/collections/adult-care-disposal", "CARE", "원래 CARE 메뉴가 가리키던 핸들"),
    ("/collections/pets사본", "/collections/medical-waste-pails", "MEDICAL", "원래 MEDICAL 메뉴가 가리키던 핸들"),
    ("/collections/care사본", "/collections/sanitary-disposal-bins", "PUBLIC", "원래 PUBLIC 메뉴가 가리키던 핸들"),
    ("/collections/office사본", "/collections/hospitality-sanitary-bins", "HOSPITALITY", "원래 HOSPITALITY 메뉴가 가리키던 핸들"),
    ("/collections/public사본", "/collections/office-sanitary-bins", "OFFICE", "원래 OFFICE 메뉴가 가리키던 핸들"),
    ("/collections/home", "/collections/home-odor-sealing-pails", "HOME", "기존 home 컬렉션을 키워드 핸들로 마이그레이션"),
    ("/pages/brand-stroy", "/pages/brand-story", "—", "Brand Story 페이지 핸들 오타 수정"),
]

for idx, (frm, to, ind, note) in enumerate(redirects, start=1):
    row = 5 + idx - 1
    ws.cell(row=row, column=1, value=idx)
    ws.cell(row=row, column=2, value=frm)
    ws.cell(row=row, column=3, value=to)
    ws.cell(row=row, column=4, value=ind)
    ws.cell(row=row, column=5, value=note)
    style_body_row(ws, row, len(headers), alt=(idx % 2 == 0))
    ws.row_dimensions[row].height = 30

# Pattern A 안내 박스
note_row = 5 + len(redirects) + 1
ws.cell(row=note_row, column=1, value="✓ Shopify 작업 순서").font = Font(name="Arial", bold=True, size=11, color=COLOR_PRIMARY)
note_row += 1
steps = [
    "1) Shopify Admin → Products → Collections → '사본' 컬렉션 클릭 → 'Search engine listing' → 'Edit' → URL handle을 위 표의 'To' 컬럼 핸들로 수정 + Title도 함께 수정 (시트 6_SEO_페이지별 참고)",
    "2) Shopify Admin → Online Store → Navigation → URL redirects → 'Add URL redirect' → 위 9개 행을 한 번씩 입력",
    "3) Shopify Admin → Online Store → Navigation → 'Main menu' 편집 → 각 메뉴 라벨(BABY/PETS/...)이 새 URL로 연결되도록 재설정",
    "4) Search Console → URL 검사 → 새 URL 9개 각각 'Request Indexing' 한 번씩 클릭",
    "5) janibell_dataLayer_snippet.liquid · janibell_gtm_container.json 두 파일은 Pattern A 매핑이 이미 적용됨 (코드 추가 작업 불필요)",
]
for s in steps:
    c = ws.cell(row=note_row, column=1, value=s)
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=5)
    c.font = FONT_BODY
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[note_row].height = 36
    note_row += 1

set_widths(ws, [5, 36, 42, 16, 36])
ws.freeze_panes = "A5"

# ===================================================================
# 마지막에 진행률 요약 시트 (선택)
# ===================================================================
ws = wb.create_sheet("10_진행률요약")
add_title(ws,
          "전체 진행률 한 눈에",
          "1_마스터체크리스트의 상태 컬럼을 자동 집계. 시트 편집 시 자동 갱신.",
          last_col=4)

ws.cell(row=4, column=1, value="카테고리"); ws.cell(row=4, column=2, value="전체"); ws.cell(row=4, column=3, value="완료"); ws.cell(row=4, column=4, value="완료율")
style_header_row(ws, 4, 4)
cats = ["A. 사전준비","B. GA4","C. GTM","D. dataLayer","E. Pixel","F. SearchConsole","G. SEO기초","H. 검증/QA","I. Looker","J. 운영루틴"]
for i, c in enumerate(cats, start=1):
    row = 4 + i
    ws.cell(row=row, column=1, value=c)
    ws.cell(row=row, column=2, value=f'=COUNTIF(\'1_마스터체크리스트\'!B:B,"{c}")')
    ws.cell(row=row, column=3, value=f'=COUNTIFS(\'1_마스터체크리스트\'!B:B,"{c}",\'1_마스터체크리스트\'!I:I,"완료")')
    ws.cell(row=row, column=4, value=f"=IFERROR(C{row}/B{row},0)")
    ws.cell(row=row, column=4).number_format = "0.0%"
    style_body_row(ws, row, 4, alt=(i % 2 == 0))

# Total row
total_row = 4 + len(cats) + 1
ws.cell(row=total_row, column=1, value="전체").font = Font(name="Arial", bold=True, size=11)
ws.cell(row=total_row, column=2, value=f"=SUM(B5:B{total_row-1})")
ws.cell(row=total_row, column=3, value=f"=SUM(C5:C{total_row-1})")
ws.cell(row=total_row, column=4, value=f"=IFERROR(C{total_row}/B{total_row},0)")
ws.cell(row=total_row, column=4).number_format = "0.0%"
style_body_row(ws, total_row, 4)
for c in range(1,5):
    ws.cell(row=total_row, column=c).fill = PatternFill("solid", fgColor=COLOR_BANNER_BG)
    ws.cell(row=total_row, column=c).font = Font(name="Arial", bold=True, size=11)

set_widths(ws, [22, 12, 12, 14])
ws.freeze_panes = "A5"

# Save
out = "/sessions/compassionate-sweet-noether/mnt/outputs/janibell_setup/Janibell_마케팅인프라_체크리스트.xlsx"
wb.save(out)
print("Saved:", out)
